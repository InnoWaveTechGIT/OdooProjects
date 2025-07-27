from odoo import models, fields, api
import base64
import xlrd
from openpyxl import load_workbook
import io
from odoo.exceptions import UserError

class ExcelImportWizard(models.TransientModel):
    _name = 'order.excel.import.wizard'
    _description = 'Excel Import Wizard'

    excel_file = fields.Binary(string='Excel File', required=True)
    order_id = fields.Many2one('nupco.orders')

    def is_data_valid(self, model_name, field_name, data):
        # Retrieve the model based on the provided model name
        Model = self.env['ir.model'].search([('name' , '=' ,model_name )])
        for i in Model.field_id:
            if i.name == field_name:
                field = i
        is_selection = field.ttype
        # Check if the field is required
        is_required = field.required
        is_selection = field.ttype
        # if is_required and data== False:
        print('data and field name >>>>>>>> ' ,data,field_name )
        if is_required and data == False:
            raise UserError("The field " + field_name  + " is invalid. the field is required and there is no data.")
        # Attempt to set the value and check if the field validation passes
        if is_selection == 'selection' and data not in field.selection_ids.mapped('value'):
            raise UserError ("The field " + field_name  + " is invalid. the field is Selection  and there is invalid data entered . it should be like" + str(field.selection_ids.mapped('value')))

        if is_selection =='many2one'  and data == None and is_required:
            if field_name == 'product_id':
                field_name = 'Trade Name'
            raise UserError("The field " + field.field_description+ " is invalid. the field is Many2one  and there is invalid data entered .")
        if field_name == 'delivery_address_id':
            print('Data  >>> ' , str(data) +' ' , str(field_name) , '  ' , str(is_selection))
        if is_selection =='many2one' and str(data) in ['False' , 'None'] and field_name in ['customer_id' , 'delivery_address_id'] :
            print('data >>>>>>>> ' , data)
            raise UserError("The field " + field.field_description+ " is invalid. the field is Many2one  and the data entered is not valid .")

        return True

    def import_excel_data(self):
        tender_obj = self.env['nubco.tender']
        uom_obj = self.env['uom.uom']
        partner = self.env['res.partner']
        currency_obj = self.env['res.currency']
        vat_obj = self.env['account.tax']
        product_obj = self.env['product.product']
        manufacturer_obj = self.env['res.partner']
        country_obj = self.env['res.country']
        packaging = self.env['product.packaging']

        for wizard in self:
            data = base64.b64decode(wizard.excel_file)
            excel_data = io.BytesIO(data)
            workbook = load_workbook(excel_data, read_only=True)
            sheet = workbook.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Create or update records based on the Excel data
                tender_obj.create({
                    'order_id': self.order_id.id,
                    'customer_id' : partner.search([('name', '=', row[2])], limit=1).id if self.is_data_valid('nubco.tender' , 'customer_id' ,partner.search([('name', '=', row[2])], limit=1).id)  else False ,
                    'customer_ref' : row[3] if row[3] != 'False' else False,
                    'etimad' : row[4] if row[4] != 'False' else False,
                    'contract_item_no' : row[5] if row[5] != 'False' else False,
                    'purchasing_document' : row[6] if row[6] != 'False' else False,
                    'po_item' : row[7] if row[7] != 'False' else False,
                    'nubco_serial' : row[8] if row[8] != 'False' else False,
                    'cust_gen_code' : row[9] if row[9] != 'False' else False,
                    'nubco_material' : row[10] if row[10] != 'False' else False,
                    'nubco_material_des' : row[11] if row[11] != 'False' else False,
                    'quantity' : row[12] if row[12] != 'False' else False,
                    'uom_id': uom_obj.search([('name', '=', row[13])], limit=1).id if self.is_data_valid('nubco.tender' , 'uom_id' ,row[13]) else False,
                    'price' : row[14] if row[14] != 'False' else False,
                    'per_unit': row[15] if row[15] != 'False' else False,
                    'net_price_per': row[16] if row[16] != 'False' else False,
                    'net_order_value' : row[17] if row[17] != 'False' else False,
                    'tax_value': row[18] if row[18] != 'False' else False,
                    'vat_id': vat_obj.search([('name', '=', row[19])], limit=1).id if row[19] != 'False' else False,
                    'delivery_address_id': partner.search([('name', '=', row[20])], limit=1).id if self.is_data_valid('nubco.tender' , 'delivery_address_id' ,partner.search([('name', '=', row[20])], limit=1).id)  else False ,
                    'nupco_delivery_date': row[21] if row[21] != 'False' else False,
                    'delivery_no': row[22] if row[22] != 'False' else False,
                    'plant': row[23] if row[23] != 'False' else False,
                    'product_id': product_obj.search([('name', '=', row[24])], limit=1).id if self.is_data_valid('nubco.tender' , 'product_id' ,row[24]) else False,
                    'manufacturer_id': manufacturer_obj.search([('name', '=', row[25])], limit=1).id if manufacturer_obj.search([('name', '=', row[25])], limit=1) else False,
                    'manufacturing_country': country_obj.search([('name', '=', row[26])], limit=1).id if row[26] != 'False' else False,
                    'product_packaging': packaging.search([('name', '=', row[27])], limit=1).id if row[27] != 'False' else False,
                    'moq': row[28] if row[28] != 'False' else False,
                    'volume': row[29] if row[29] != 'False' else False,
                    'manufacture_process_local': row[30] if self.is_data_valid('nubco.tender' , 'manufacture_process_local' ,row[30]) else False,
                    'temp': row[31] if row[31] != 'False' else False,
                    'first_lead': row[32] if row[32] != 'False' else False,
                    'lead_time': row[33] if row[33] != 'False' else False,
                    'max_num_of_shipp': row[34] if row[34] != 'False' else False,
                    'item_type' : 'Orig',
                    'official_agent' : 'yes'

            })

        return {'type': 'ir.actions.act_window_close'}
