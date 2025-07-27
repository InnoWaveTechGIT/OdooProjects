from odoo import models, fields, api ,_
import base64
import xlrd
from openpyxl import load_workbook
import io
from odoo.exceptions import UserError

class ExcelImportWizard(models.TransientModel):
    _name = 'excel.import.wizard'
    _description = 'Excel Import Wizard'

    excel_file = fields.Binary(string='Excel File', required=True)
    tender_id = fields.Many2one('tenders')

    def is_number(self ,value):
        try:
            # Attempt to convert the value to an integer
            int_value = int(value)
            return True
        except (ValueError, TypeError):
            return False
    def is_data_valid(self, model_name, field_name, data):
        # Retrieve the model based on the provided model name
        Model = self.env['ir.model'].search([('name' , '=' ,model_name )])
        for i in Model.field_id:
            if i.name == field_name:
                field = i
        is_selection = field.ttype
        # Check if the field is required
        is_required = field.required
        is_selection = field.ttype
        # if is_required and data== False:
        if is_required and data == False:
            raise UserError("The field " + field_name  + " is invalid. the field is required and there is no data.")
        # Attempt to set the value and check if the field validation passes
        if is_selection == 'selection' and data not in field.selection_ids.mapped('value'):
            raise UserError ("The field " + field_name  + " is invalid. the field is Selection  and there is invalid data entered . it should be like" + str(field.selection_ids.mapped('value')))

        if is_selection =='many2one'  and data == None and is_required:
            if field_name == 'product_id':
                field_name = 'Trade Name'
            raise UserError("The field " + field.field_description+ " is invalid. the field is Many2one  and there is invalid data entered .")

        return True

    def import_excel_data(self):
        tender_obj = self.env['nubco.tender']
        uom_obj = self.env['uom.uom']
        currency_obj = self.env['res.currency']
        vat_obj = self.env['account.tax']
        product_obj = self.env['product.product']
        manufacturer_obj = self.env['res.partner']
        country_obj = self.env['res.country']

        for wizard in self:
            data = base64.b64decode(wizard.excel_file)
            excel_data = io.BytesIO(data)
            workbook = load_workbook(excel_data, read_only=True)
            sheet = workbook.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Create or update records based on the Excel data
                tender_line = tender_obj.create({
                'tender_id': self.tender_id.id,
                'change_option' : row[2] if self.is_data_valid('nubco.tender' , 'change_option' ,row[2]) else False,
                'item_type' : row[3] if self.is_data_valid('nubco.tender' , 'item_type' ,row[3])== True else False,
                'nubco_serial': row[4] if self.is_data_valid('nubco.tender' , 'nubco_serial' ,row[4]) else False,
                'nubco_material': row[5] if self.is_data_valid('nubco.tender' , 'nubco_material' ,row[5]) else False,
                'nubco_material_des': row[6] if self.is_data_valid('nubco.tender' , 'nubco_material_des' ,row[6]) else False,
                    'medical_group': row[7] if self.is_data_valid('nubco.tender' , 'medical_group' ,row[7]) else False,
                    'quantity1': row[8] if self.is_data_valid('nubco.tender' , 'quantity1' ,row[8]) else False,
                     'uom_id': uom_obj.search([('name', '=', row[9])], limit=1).id if self.is_data_valid('nubco.tender' , 'uom_id' ,row[9]) else False ,
                    'price': row[10] if self.is_data_valid('nubco.tender' , 'price' ,row[10]) else False,
                  'currency_id': currency_obj.search([('name', '=', row[11])], limit=1).id if currency_obj.search([('name', '=', row[11])], limit=1) else False ,
                 'vat_id': vat_obj.search([('name', 'ilike', row[12])], limit=1).id if vat_obj.search([('name', '=', row[12])], limit=1) else False,
                'product_id': product_obj.search([('name', 'ilike', row[13])], limit=1).id if self.is_data_valid('nubco.tender' , 'product_id' ,row[13]) else False ,
                'manufacturer_id': manufacturer_obj.search([('name', '=', row[14])], limit=1).id if manufacturer_obj.search([('name', '=', row[14])], limit=1) else False,
                'manufacturing_country': country_obj.search([('name', '=', row[15])], limit=1).id if country_obj.search([('name', '=', row[15])], limit=1) else False,
                'product_packaging': row[16] if self.is_data_valid('nubco.tender' , 'product_packaging' ,row[16]) else False,
                'mdma_code' : row[17] if self.is_data_valid('nubco.tender' , 'mdma_code' ,row[17]) else False,
                'foc' : row[18] if self.is_data_valid('nubco.tender' , 'foc' ,row[18]) else False,
                'MDMA_exp' : row[19] if self.is_data_valid('nubco.tender' , 'MDMA_exp' ,row[19]) else False,
                'SFGa_code' : row[20] if self.is_data_valid('nubco.tender' , 'SFGa_code' ,row[20]) else False,
                'concentration' : row[21] if self.is_data_valid('nubco.tender' , 'concentration' ,row[21]) else False,
                'sheif_life' : row[22] if self.is_data_valid('nubco.tender' , 'sheif_life' ,row[22]) else False,
                'moq': row[23] if self.is_data_valid('nubco.tender' , 'moq' ,row[24]) else False,
                'volume': row[24] if self.is_data_valid('nubco.tender' , 'nubco_material' ,row[5]) else False,
                'manufacture_process_local': row[25] if self.is_data_valid('nubco.tender' , 'manufacture_process_local' ,row[25]) else False,
                'official_agent': row[26] if self.is_data_valid('nubco.tender' , 'official_agent' ,row[26]) else False,
                'company_comment': row[27] if self.is_data_valid('nubco.tender' , 'company_comment' ,row[27]) else False,
                'company_comment_term': row[28] if self.is_data_valid('nubco.tender' , 'company_comment_term' ,row[28]) else False,
                'temp': row[29] if self.is_data_valid('nubco.tender' , 'temp' ,row[29]) else False,
                'first_lead': row[30] if self.is_data_valid('nubco.tender' , 'first_lead' ,row[30]) else False,
                'lead_time': row[31] if self.is_data_valid('nubco.tender' , 'lead_time' ,row[31]) else False,
                'max_num_of_shipp': row[32] if self.is_data_valid('nubco.tender' , 'max_num_of_shipp' ,row[32]) else False,
            })

        # except Exception as e:
        #     raise UserError("The file is invalid. Please check the file format and contents.")

        return {'type': 'ir.actions.act_window_close'}
